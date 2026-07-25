"""
Генерация путевого листа форма №4-с (грузовой автомобиль, Постановление
Госкомстата России от 28.11.1997 №78) в форматах Excel и PDF.

Отличия от формы №3 (см. app/forms.py): прицепы (до 2), грузовые
реквизиты — пункты погрузки/разгрузки, номера ТТН, грузоотправитель/
грузополучатель. Та же MVP-оговорка: полный набор обязательных реквизитов
из ТЗ, упрощённая вёрстка (не факсимиле бланка Госкомстата).
"""
from __future__ import annotations

from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Paragraph

from .forms import _FONT_BOLD, _FONT_OBLIQUE, _FONT_REGULAR, _PDF_STYLE
from .schemas import ЗапросФормы4С, ПутевойЛист, РезультатРасчёта

_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE_FONT = Font(bold=True, size=13)
_LABEL_FONT = Font(bold=True, size=9, color="555555")
_VALUE_FONT = Font(size=11)


# ------- Excel -------


def _put(ws: Worksheet, row: int, label: str, value: object, col_label: int = 1, col_value: int = 2) -> None:
    lc = ws.cell(row=row, column=col_label, value=label)
    lc.font = _LABEL_FONT
    vc = ws.cell(row=row, column=col_value, value=value if value not in (None, "") else "—")
    vc.font = _VALUE_FONT
    vc.border = _BORDER


def _sheet_for_list(wb: Workbook, req: ЗапросФормы4С, лист: ПутевойЛист) -> Worksheet:
    ws = wb.create_sheet(title=f"ПЛ №{лист.номер}"[:31])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 45

    row = 1
    ws.cell(row=row, column=1, value="ПУТЕВОЙ ЛИСТ ГРУЗОВОГО АВТОМОБИЛЯ (форма №4-с)").font = _TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    ws.cell(row=row, column=1, value=f"№ {лист.номер}").font = _VALUE_FONT
    row += 2

    поля = req.поля

    if поля.организация:
        ws.cell(row=row, column=1, value="Организация").font = Font(bold=True, size=11)
        row += 1
        org = req.организация
        _put(ws, row, "Наименование", org.наименование); row += 1
        _put(ws, row, "ИНН", org.инн); row += 1
        _put(ws, row, "Адрес", org.адрес); row += 1
        _put(ws, row, "Телефон", org.телефон); row += 1
        row += 1

    if поля.тс:
        ws.cell(row=row, column=1, value="Транспортное средство").font = Font(bold=True, size=11)
        row += 1
        marka_model = " ".join(x for x in [req.расчёт.марка, req.расчёт.модель] if x)
        _put(ws, row, "Марка/модель", marka_model); row += 1
        _put(ws, row, "Тип", req.тс.тип or "грузовой"); row += 1
        _put(ws, row, "Гос. номер", req.тс.госномер); row += 1
        if req.тс.гаражныйНомер:
            _put(ws, row, "Гаражный номер", req.тс.гаражныйНомер); row += 1
        for i, пр in enumerate(req.прицепы, start=1):
            _put(ws, row, f"Прицеп {i} (марка/модель)", пр.маркаМодель); row += 1
            _put(ws, row, f"Прицеп {i} (гос. номер)", пр.госномер); row += 1
        row += 1

    if поля.водитель:
        ws.cell(row=row, column=1, value="Водитель").font = Font(bold=True, size=11)
        row += 1
        drv_doc = req.водитель
        фио = drv_doc.фио or лист.водитель
        _put(ws, row, "ФИО", фио); row += 1
        _put(ws, row, "Водительское удостоверение", drv_doc.удостоверение); row += 1
        _put(ws, row, "Класс ТС", drv_doc.классТС); row += 1
        _put(ws, row, "СНИЛС", drv_doc.снилс); row += 1
        row += 1

    ws.cell(row=row, column=1, value="Тип перевозки").font = Font(bold=True, size=11)
    row += 1
    _put(ws, row, "Тип перевозки", req.типПеревозки.value); row += 1
    _put(ws, row, "Вид сообщения", лист.видСообщения.value); row += 1
    row += 1

    if поля.показанияОдометра or поля.гсм:
        ws.cell(row=row, column=1, value="Работа автомобиля и ГСМ").font = Font(bold=True, size=11)
        row += 1
        _put(ws, row, "Дата/время выпуска на линию", лист.выпуск); row += 1
        _put(ws, row, "Дата/время возвращения", лист.возвращение); row += 1
        _put(ws, row, "Общее время за выезд, ч", лист.общееВремя); row += 1
        if поля.показанияОдометра:
            _put(ws, row, "Одометр на выдачу, км", лист.одометрВыдача); row += 1
            _put(ws, row, "Одометр на закрытие, км", лист.одометрЗакрытие); row += 1
            _put(ws, row, "Общий пробег, км", лист.пробег); row += 1
        if поля.гсм:
            _put(ws, row, "Остаток ГСМ на выдачу, л", лист.остатокВыдача); row += 1
            _put(ws, row, "Остаток ГСМ на закрытие, л", лист.остатокЗакрытие); row += 1
            _put(ws, row, "Расход ГСМ норма, л", лист.расходНорма); row += 1
            _put(ws, row, "Расход ГСМ факт, л", лист.расходФакт); row += 1
        row += 1

    if поля.груз and req.ездки:
        ws.cell(row=row, column=1, value="Задание водителю (груз)").font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value="№").font = _LABEL_FONT
        ws.cell(row=row, column=2, value="Погрузка → Разгрузка / Груз / ТТН").font = _LABEL_FONT
        row += 1
        for i, ездка in enumerate(req.ездки, start=1):
            summary = (
                f"{ездка.пунктПогрузки or '—'} → {ездка.пунктРазгрузки or '—'} | "
                f"груз: {ездка.наименованиеГруза or '—'} | "
                f"ТТН №{ездка.номерТТН or '—'} | "
                f"грузоотправитель: {ездка.грузоотправитель or '—'} | "
                f"грузополучатель: {ездка.грузополучатель or '—'}"
                + (f" | {ездка.весТонн} т" if ездка.весТонн else "")
            )
            ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=2, value=summary).font = _VALUE_FONT
            row += 1
        row += 1

    if поля.маршрут and лист.маршрут:
        ws.cell(row=row, column=1, value="Маршрут").font = Font(bold=True, size=11)
        row += 1
        for i, stop in enumerate(лист.маршрут, start=1):
            ws.cell(row=row, column=1, value=f"{i}.").alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=2, value=stop).font = _VALUE_FONT
            row += 1
        row += 1

    row += 1
    disclaimer = (
        "Внимание! Путевой лист без отметок о прохождении медицинского осмотра "
        "водителя и технического контроля транспортного средства недействителен."
    )
    ws.cell(row=row, column=1, value=disclaimer).font = Font(italic=True, size=8, color="777777")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

    return ws


def generate_form4c_excel(req: ЗапросФормы4С, result: РезультатРасчёта) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if not result.листы:
        ws = wb.create_sheet(title="Нет данных")
        ws.cell(row=1, column=1, value="Путевые листы не сформированы — см. предупреждения расчёта.")
    else:
        for лист in result.листы:
            _sheet_for_list(wb, req, лист)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------- PDF -------


def _pdf_page(c: pdf_canvas.Canvas, req: ЗапросФормы4С, лист: ПутевойЛист) -> None:
    width, height = A4
    x_left = 20 * mm
    x_right = width - 20 * mm
    y = height - 20 * mm

    def line(label: str, value: object, gap: float = 6 * mm) -> None:
        nonlocal y
        c.setFont(_FONT_BOLD, 8)
        c.setFillColor(colors.grey)
        c.drawString(x_left, y, label)
        c.setFont(_FONT_REGULAR, 10)
        c.setFillColor(colors.black)
        c.drawString(x_left + 70 * mm, y, str(value) if value not in (None, "") else "—")
        y -= gap

    def heading(text: str) -> None:
        nonlocal y
        y -= 3 * mm
        c.setFont(_FONT_BOLD, 11)
        c.setFillColor(colors.black)
        c.drawString(x_left, y, text)
        y -= 6 * mm

    c.setFont(_FONT_BOLD, 14)
    c.drawCentredString(width / 2, y, "ПУТЕВОЙ ЛИСТ ГРУЗОВОГО АВТОМОБИЛЯ (форма №4-с)")
    y -= 8 * mm
    c.setFont(_FONT_REGULAR, 11)
    c.drawCentredString(width / 2, y, f"№ {лист.номер}")
    y -= 10 * mm
    c.line(x_left, y, x_right, y)
    y -= 6 * mm

    поля = req.поля

    if поля.организация:
        heading("Организация")
        org = req.организация
        line("Наименование", org.наименование)
        line("ИНН", org.инн)
        line("Адрес", org.адрес)
        line("Телефон", org.телефон)

    if поля.тс:
        heading("Транспортное средство")
        marka_model = " ".join(x for x in [req.расчёт.марка, req.расчёт.модель] if x)
        line("Марка/модель", marka_model)
        line("Гос. номер", req.тс.госномер)
        for i, пр in enumerate(req.прицепы, start=1):
            line(f"Прицеп {i}", f"{пр.маркаМодель} {пр.госномер}".strip())

    if поля.водитель:
        heading("Водитель")
        drv_doc = req.водитель
        line("ФИО", drv_doc.фио or лист.водитель)
        line("Водительское удостоверение", drv_doc.удостоверение)
        line("Класс ТС", drv_doc.классТС)
        line("СНИЛС", drv_doc.снилс)

    heading("Тип перевозки / сообщения")
    line("Тип перевозки", req.типПеревозки.value)
    line("Вид сообщения", лист.видСообщения.value)

    if поля.показанияОдометра or поля.гсм:
        heading("Работа автомобиля и ГСМ")
        line("Выпуск на линию", лист.выпуск)
        line("Возвращение", лист.возвращение)
        line("Общее время, ч", лист.общееВремя)
        if поля.показанияОдометра:
            line("Одометр на выдачу, км", лист.одометрВыдача)
            line("Одометр на закрытие, км", лист.одометрЗакрытие)
            line("Общий пробег, км", лист.пробег)
        if поля.гсм:
            line("Остаток ГСМ на выдачу, л", лист.остатокВыдача)
            line("Остаток ГСМ на закрытие, л", лист.остатокЗакрытие)
            line("Расход норма/факт, л", f"{лист.расходНорма} / {лист.расходФакт}")

    if поля.груз and req.ездки:
        heading("Задание водителю (груз)")
        for i, ездка in enumerate(req.ездки, start=1):
            summary = (
                f"{i}. {ездка.пунктПогрузки or '—'} → {ездка.пунктРазгрузки or '—'} | "
                f"груз: {ездка.наименованиеГруза or '—'} | ТТН №{ездка.номерТТН or '—'} | "
                f"грузоотправитель: {ездка.грузоотправитель or '—'} | "
                f"грузополучатель: {ездка.грузополучатель or '—'}"
                + (f" | {ездка.весТонн} т" if ездка.весТонн else "")
            )
            p = Paragraph(summary, _PDF_STYLE)
            w, h = p.wrap(x_right - x_left, 20 * mm)
            p.drawOn(c, x_left, y - h + 10)
            y -= h + 2 * mm

    if поля.маршрут and лист.маршрут:
        heading("Маршрут")
        for i, stop in enumerate(лист.маршрут, start=1):
            p = Paragraph(f"{i}. {stop}", _PDF_STYLE)
            w, h = p.wrap(x_right - x_left, 20 * mm)
            p.drawOn(c, x_left, y - h + 10)
            y -= h + 2 * mm

    y -= 8 * mm
    c.setFont(_FONT_OBLIQUE, 7)
    c.setFillColor(colors.grey)
    disclaimer = (
        "Внимание! Путевой лист без отметок о прохождении медицинского осмотра водителя "
        "и технического контроля транспортного средства недействителен."
    )
    c.drawString(x_left, max(y, 15 * mm), disclaimer)

    c.showPage()


def generate_form4c_pdf(req: ЗапросФормы4С, result: РезультатРасчёта) -> bytes:
    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)

    if not result.листы:
        width, height = A4
        c.setFont(_FONT_REGULAR, 12)
        c.drawCentredString(width / 2, height / 2, "Путевые листы не сформированы.")
        c.showPage()
    else:
        for лист in result.листы:
            _pdf_page(c, req, лист)

    c.save()
    return buf.getvalue()
