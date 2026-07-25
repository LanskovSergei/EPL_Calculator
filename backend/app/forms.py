"""
Генерация путевого листа форма №3 (легковой автомобиль, Постановление
Госкомстата России от 28.11.1997 №78) в форматах Excel и PDF.

MVP-раскладка: не пиксель-в-пиксель копия бланка Госкомстата, а полный
набор обязательных реквизитов из ТЗ (раздел «В каждом путевом листе
указываются»), сгруппированный по смысловым блокам на одной странице/листе
на каждый путевой лист. Визуальное приближение к официальному бланку —
отдельная доработка (брендирование/вёрстка, вне этого шага).

Один путевой лист = один лист Excel / одна страница PDF (ТЗ: «Один путевой
лист выдаётся на одну смену одного водителя на одном ТС»).
"""
from __future__ import annotations

import os
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Paragraph

from .schemas import ЗапросФормыПЛ, ПутевойЛист, РезультатРасчёта

# reportlab base-14 (Helvetica и т.п.) не содержат кириллицу — регистрируем
# DejaVu Sans (шрифт с кириллицей, идёт в комплекте с проектом).
_FONTS_DIR = os.path.join(os.path.dirname(__file__), "fonts")
_FONT_REGULAR = "DejaVuSans"
_FONT_BOLD = "DejaVuSans-Bold"
_FONT_OBLIQUE = "DejaVuSans-Oblique"

if _FONT_REGULAR not in pdfmetrics.getRegisteredFontNames():
    pdfmetrics.registerFont(TTFont(_FONT_REGULAR, os.path.join(_FONTS_DIR, "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont(_FONT_BOLD, os.path.join(_FONTS_DIR, "DejaVuSans-Bold.ttf")))
    pdfmetrics.registerFont(TTFont(_FONT_OBLIQUE, os.path.join(_FONTS_DIR, "DejaVuSans-Oblique.ttf")))


# ------- Excel -------

_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE_FONT = Font(bold=True, size=13)
_LABEL_FONT = Font(bold=True, size=9, color="555555")
_VALUE_FONT = Font(size=11)


def _put(ws: Worksheet, row: int, label: str, value: object, col_label: int = 1, col_value: int = 2) -> None:
    lc = ws.cell(row=row, column=col_label, value=label)
    lc.font = _LABEL_FONT
    vc = ws.cell(row=row, column=col_value, value=value if value not in (None, "") else "—")
    vc.font = _VALUE_FONT
    vc.border = _BORDER


def _sheet_for_list(wb: Workbook, req: ЗапросФормыПЛ, лист: ПутевойЛист) -> Worksheet:
    ws = wb.create_sheet(title=f"ПЛ №{лист.номер}"[:31])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45

    row = 1
    ws.cell(row=row, column=1, value="ПУТЕВОЙ ЛИСТ ЛЕГКОВОГО АВТОМОБИЛЯ (форма №3)").font = _TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    ws.cell(row=row, column=1, value=f"№ {лист.номер}").font = _VALUE_FONT
    row += 2

    поля = req.поля

    if поля.организация:
        ws.cell(row=row, column=1, value="Организация").font = Font(bold=True, size=11)
        row += 1
        org = req.организация
        _put(ws, row, "Наименование", org.наименование); row += 1
        _put(ws, row, "ИНН", org.инн); row += 1
        _put(ws, row, "Адрес", org.адрес); row += 1
        _put(ws, row, "Телефон", org.телефон); row += 1
        if org.окпо:
            _put(ws, row, "ОКПО", org.окпо); row += 1
        row += 1

    if поля.тс:
        ws.cell(row=row, column=1, value="Транспортное средство").font = Font(bold=True, size=11)
        row += 1
        marka_model = " ".join(x for x in [req.расчёт.марка, req.расчёт.модель] if x)
        _put(ws, row, "Марка/модель", marka_model); row += 1
        _put(ws, row, "Тип", req.тс.тип or "легковой"); row += 1
        _put(ws, row, "Гос. номер", req.тс.госномер); row += 1
        if req.тс.гаражныйНомер:
            _put(ws, row, "Гаражный номер", req.тс.гаражныйНомер); row += 1
        if req.прицеп:
            _put(ws, row, "Прицеп (марка/модель)", req.прицеп.маркаМодель); row += 1
            _put(ws, row, "Прицеп (гос. номер)", req.прицеп.госномер); row += 1
        row += 1

    if поля.водитель:
        ws.cell(row=row, column=1, value="Водитель").font = Font(bold=True, size=11)
        row += 1
        drv_doc = req.водитель
        фио = drv_doc.фио or лист.водитель
        _put(ws, row, "ФИО", фио); row += 1
        _put(ws, row, "Водительское удостоверение", drv_doc.удостоверение); row += 1
        _put(
            ws, row, "Дата выдачи ВУ",
            drv_doc.датаВыдачиУдостоверения.isoformat() if drv_doc.датаВыдачиУдостоверения else "",
        )
        row += 1
        _put(ws, row, "Класс ТС", drv_doc.классТС); row += 1
        _put(ws, row, "СНИЛС", drv_doc.снилс); row += 1
        row += 1

    ws.cell(row=row, column=1, value="Тип перевозки").font = Font(bold=True, size=11)
    row += 1
    _put(ws, row, "Тип перевозки", req.типПеревозки.value); row += 1
    _put(ws, row, "Вид сообщения", лист.видСообщения.value); row += 1
    row += 1

    if поля.показанияОдометра or поля.гсм:
        ws.cell(row=row, column=1, value="Работа автомобиля и ГСМ").font = Font(bold=True, size=11)
        row += 1
        _put(ws, row, "Дата/время выпуска на линию", лист.выпуск); row += 1
        _put(ws, row, "Дата/время возвращения", лист.возвращение); row += 1
        _put(ws, row, "Общее время за выезд, ч", лист.общееВремя); row += 1
        if поля.показанияОдометра:
            _put(ws, row, "Одометр на выдачу, км", лист.одометрВыдача); row += 1
            _put(ws, row, "Одометр на закрытие, км", лист.одометрЗакрытие); row += 1
            _put(ws, row, "Общий пробег, км", лист.пробег); row += 1
        if поля.гсм:
            _put(ws, row, "Остаток ГСМ на выдачу, л", лист.остатокВыдача); row += 1
            _put(ws, row, "Остаток ГСМ на закрытие, л", лист.остатокЗакрытие); row += 1
            _put(ws, row, "Расход ГСМ норма, л", лист.расходНорма); row += 1
            _put(ws, row, "Расход ГСМ факт, л", лист.расходФакт); row += 1
        row += 1

    if поля.маршрут and лист.маршрут:
        ws.cell(row=row, column=1, value="Маршрут").font = Font(bold=True, size=11)
        row += 1
        for i, stop in enumerate(лист.маршрут, start=1):
            ws.cell(row=row, column=1, value=f"{i}.").alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=2, value=stop).font = _VALUE_FONT
            row += 1
        row += 1

    row += 1
    disclaimer = (
        "Внимание! Путевой лист без отметок о прохождении медицинского осмотра "
        "водителя и технического контроля транспортного средства недействителен."
    )
    ws.cell(row=row, column=1, value=disclaimer).font = Font(italic=True, size=8, color="777777")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

    return ws


def generate_form3_excel(req: ЗапросФормыПЛ, result: РезультатРасчёта) -> bytes:
    """Одна книга Excel, один лист на каждый путевой лист."""
    wb = Workbook()
    wb.remove(wb.active)  # убираем дефолтный пустой лист

    if not result.листы:
        ws = wb.create_sheet(title="Нет данных")
        ws.cell(row=1, column=1, value="Путевые листы не сформированы — см. предупреждения расчёта.")
    else:
        for лист in result.листы:
            _sheet_for_list(wb, req, лист)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------- PDF -------

_PDF_STYLE = ParagraphStyle(name="body", fontName=_FONT_REGULAR, fontSize=9, leading=12)
_PDF_LABEL_STYLE = ParagraphStyle(name="label", fontName=_FONT_BOLD, fontSize=8, textColor=colors.grey)


def _pdf_page(c: pdf_canvas.Canvas, req: ЗапросФормыПЛ, лист: ПутевойЛист) -> None:
    width, height = A4
    x_left = 20 * mm
    x_right = width - 20 * mm
    y = height - 20 * mm

    def line(label: str, value: object, gap: float = 6 * mm) -> None:
        nonlocal y
        c.setFont(_FONT_BOLD, 8)
        c.setFillColor(colors.grey)
        c.drawString(x_left, y, label)
        c.setFont(_FONT_REGULAR, 10)
        c.setFillColor(colors.black)
        c.drawString(x_left + 70 * mm, y, str(value) if value not in (None, "") else "—")
        y -= gap

    def heading(text: str) -> None:
        nonlocal y
        y -= 3 * mm
        c.setFont(_FONT_BOLD, 11)
        c.setFillColor(colors.black)
        c.drawString(x_left, y, text)
        y -= 6 * mm

    c.setFont(_FONT_BOLD, 14)
    c.drawCentredString(width / 2, y, "ПУТЕВОЙ ЛИСТ ЛЕГКОВОГО АВТОМОБИЛЯ (форма №3)")
    y -= 8 * mm
    c.setFont(_FONT_REGULAR, 11)
    c.drawCentredString(width / 2, y, f"№ {лист.номер}")
    y -= 10 * mm
    c.line(x_left, y, x_right, y)
    y -= 6 * mm

    поля = req.поля

    if поля.организация:
        heading("Организация")
        org = req.организация
        line("Наименование", org.наименование)
        line("ИНН", org.инн)
        line("Адрес", org.адрес)
        line("Телефон", org.телефон)

    if поля.тс:
        heading("Транспортное средство")
        marka_model = " ".join(x for x in [req.расчёт.марка, req.расчёт.модель] if x)
        line("Марка/модель", marka_model)
        line("Гос. номер", req.тс.госномер)
        if req.прицеп:
            line("Прицеп", f"{req.прицеп.маркаМодель} {req.прицеп.госномер}".strip())

    if поля.водитель:
        heading("Водитель")
        drv_doc = req.водитель
        line("ФИО", drv_doc.фио or лист.водитель)
        line("Водительское удостоверение", drv_doc.удостоверение)
        line("Класс ТС", drv_doc.классТС)
        line("СНИЛС", drv_doc.снилс)

    heading("Тип перевозки / сообщения")
    line("Тип перевозки", req.типПеревозки.value)
    line("Вид сообщения", лист.видСообщения.value)

    if поля.показанияОдометра or поля.гсм:
        heading("Работа автомобиля и ГСМ")
        line("Выпуск на линию", лист.выпуск)
        line("Возвращение", лист.возвращение)
        line("Общее время, ч", лист.общееВремя)
        if поля.показанияОдометра:
            line("Одометр на выдачу, км", лист.одометрВыдача)
            line("Одометр на закрытие, км", лист.одометрЗакрытие)
            line("Общий пробег, км", лист.пробег)
        if поля.гсм:
            line("Остаток ГСМ на выдачу, л", лист.остатокВыдача)
            line("Остаток ГСМ на закрытие, л", лист.остатокЗакрытие)
            line("Расход норма/факт, л", f"{лист.расходНорма} / {лист.расходФакт}")

    if поля.маршрут and лист.маршрут:
        heading("Маршрут")
        for i, stop in enumerate(лист.маршрут, start=1):
            p = Paragraph(f"{i}. {stop}", _PDF_STYLE)
            w, h = p.wrap(x_right - x_left, 20 * mm)
            p.drawOn(c, x_left, y - h + 10)
            y -= h + 2 * mm

    y -= 8 * mm
    c.setFont(_FONT_OBLIQUE, 7)
    c.setFillColor(colors.grey)
    disclaimer = (
        "Внимание! Путевой лист без отметок о прохождении медицинского осмотра водителя "
        "и технического контроля транспортного средства недействителен."
    )
    c.drawString(x_left, max(y, 15 * mm), disclaimer)

    c.showPage()


def generate_form3_pdf(req: ЗапросФормыПЛ, result: РезультатРасчёта) -> bytes:
    """Один PDF, одна страница на каждый путевой лист."""
    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)

    if not result.листы:
        width, height = A4
        c.setFont(_FONT_REGULAR, 12)
        c.drawCentredString(width / 2, height / 2, "Путевые листы не сформированы.")
        c.showPage()
    else:
        for лист in result.листы:
            _pdf_page(c, req, лист)

    c.save()
    return buf.getvalue()
