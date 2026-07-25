"""Тесты app/forms_4c.py (генерация путевого листа форма №4-с, грузовой)."""
from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app.calc import calculate
from app.forms_4c import generate_form4c_excel, generate_form4c_pdf
from app.schemas import (
    Водитель,
    Заправка,
    ВидСообщения,
    ВидТоплива,
    ТипТС,
    ВходныеДанные,
    ЗапросФормы4С,
    ДанныеОрганизации,
    ДанныеТСДокумент,
    ДанныеВодителяДокумент,
    ДанныеПрицепа,
    ЕздкаГруза,
)


def make_calc_input(**overrides) -> ВходныеДанные:
    base = dict(
        марка="Hyundai",
        модель="Porter 2",
        типТС=ТипТС.грузовой,
        видТоплива=ВидТоплива.дизель,
        объёмБака=100,
        старше10лет=False,
        прицепГруз=True,
        спецтехника=False,
        периодС=date(2025, 6, 1),
        периодПо=date(2025, 6, 30),
        видСообщения=ВидСообщения.городское,
        одометрНаНачало=32000,
        остатокНаНачало=15,
        адресСтоянки="г. Москва, ул. Ленина, 1",
        водители=[Водитель(фио="Иванов И.И.", дни=[date(2025, 6, 2)])],
        заправки=[
            Заправка(дата=date(2025, 6, 1), время="08:00", объём=40, адрес="г. Москва, АЗС Лукойл")
        ],
    )
    base.update(overrides)
    return ВходныеДанные(**base)


def make_form_request(**overrides) -> ЗапросФормы4С:
    base = dict(
        расчёт=make_calc_input(),
        организация=ДанныеОрганизации(наименование='ООО "Дубрава"', инн="7701234567"),
        тс=ДанныеТСДокумент(тип="грузовой фургон", госномер="А900ТТ178"),
        водитель=ДанныеВодителяДокумент(фио="Иванов И.И.", удостоверение="78 АБ 654321"),
        прицепы=[ДанныеПрицепа(маркаМодель="СЗАП 8357", госномер="АК123178")],
        ездки=[
            ЕздкаГруза(
                пунктПогрузки="г. Санкт-Петербург, ул. Заводская, 12",
                пунктРазгрузки="г. Санкт-Петербург, ул. Весенняя, 58",
                наименованиеГруза="Мебель",
                номерТТН="9332",
                грузоотправитель='ООО "Дубрава"',
                грузополучатель='ООО "Ремонт"',
                весТонн=0.1,
            )
        ],
    )
    base.update(overrides)
    return ЗапросФормы4С(**base)


def test_form4c_rejects_легковой():
    with pytest.raises(Exception):
        make_form_request(расчёт=make_calc_input(типТС=ТипТС.легковой))


def test_form4c_max_two_prицепа():
    with pytest.raises(Exception):
        make_form_request(
            прицепы=[
                ДанныеПрицепа(маркаМодель="A", госномер="1"),
                ДанныеПрицепа(маркаМодель="B", госномер="2"),
                ДанныеПрицепа(маркаМодель="C", госномер="3"),
            ]
        )


def test_excel_has_one_sheet_per_list():
    req = make_form_request()
    result = calculate(req.расчёт)
    content = generate_form4c_excel(req, result)
    wb = load_workbook(BytesIO(content))
    assert len(wb.sheetnames) == len(result.листы)
    assert len(result.листы) >= 1


def test_excel_contains_cargo_and_trailer_fields():
    req = make_form_request()
    result = calculate(req.расчёт)
    content = generate_form4c_excel(req, result)
    wb = load_workbook(BytesIO(content))
    ws = wb[wb.sheetnames[0]]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "Дубрава" in text
    assert "СЗАП" in text
    assert "9332" in text  # номер ТТН
    assert "Мебель" in text
    assert "форма №4-с" in text.lower() or "4-с" in text


def test_pdf_has_one_page_per_list():
    req = make_form_request()
    result = calculate(req.расчёт)
    content = generate_form4c_pdf(req, result)
    reader = PdfReader(BytesIO(content))
    assert len(reader.pages) == len(result.листы)


def test_pdf_contains_ttn_and_trailer():
    req = make_form_request()
    result = calculate(req.расчёт)
    content = generate_form4c_pdf(req, result)
    reader = PdfReader(BytesIO(content))
    text = "\n".join(p.extract_text() or "" for p in reader.pages)
    assert "9332" in text
    assert "АК123178" in text


def test_no_lists_returns_placeholder():
    req = make_form_request(
        расчёт=make_calc_input(водители=[Водитель(фио="А", дни=[])])
    )
    result = calculate(req.расчёт)
    assert result.листы == []
    content = generate_form4c_excel(req, result)
    wb = load_workbook(BytesIO(content))
    assert len(wb.sheetnames) == 1
