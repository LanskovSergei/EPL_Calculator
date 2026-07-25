"""Тесты app/forms.py (генерация путевого листа форма №3) + эндпоинтов."""
from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app.calc import calculate
from app.forms import generate_form3_excel, generate_form3_pdf
from app.schemas import (
    Водитель,
    Заправка,
    ВидСообщения,
    ВидТоплива,
    ТипТС,
    ВходныеДанные,
    ЗапросФормыПЛ,
    ДанныеОрганизации,
    ДанныеТСДокумент,
    ДанныеВодителяДокумент,
    ДанныеПрицепа,
)


def make_calc_input(**overrides) -> ВходныеДанные:
    base = dict(
        марка="Toyota",
        модель="Camry",
        типТС=ТипТС.легковой,
        видТоплива=ВидТоплива.бензин,
        объёмБака=60,
        старше10лет=False,
        прицепГруз=False,
        спецтехника=False,
        периодС=date(2025, 6, 1),
        периодПо=date(2025, 6, 30),
        видСообщения=ВидСообщения.городское,
        одометрНаНачало=50000,
        остатокНаНачало=20,
        адресСтоянки="г. Москва, ул. Ленина, 1",
        водители=[Водитель(фио="Сидоров С.С.", дни=[date(2025, 6, 2)])],
        заправки=[
            Заправка(
                дата=date(2025, 6, 1), время="08:00", объём=30, адрес="г. Москва, АЗС Роснефть"
            )
        ],
    )
    base.update(overrides)
    return ВходныеДанные(**base)


def make_form_request(**overrides) -> ЗапросФормыПЛ:
    base = dict(
        расчёт=make_calc_input(),
        организация=ДанныеОрганизации(
            наименование='ООО "Ромашка"', инн="7701234567", адрес="г. Москва", телефон="+7 495 000-00-00"
        ),
        тс=ДанныеТСДокумент(тип="легковой служебный", госномер="А123ВС777"),
        водитель=ДанныеВодителяДокумент(
            фио="Сидоров С.С.", удостоверение="77 АБ 123456", классТС="B"
        ),
    )
    base.update(overrides)
    return ЗапросФормыПЛ(**base)


# ------- Валидация -------


def test_form3_rejects_грузовой():
    with pytest.raises(Exception):
        make_form_request(расчёт=make_calc_input(типТС=ТипТС.грузовой))


# ------- Excel -------


def test_excel_has_one_sheet_per_list():
    req = make_form_request()
    result = calculate(req.расчёт)
    content = generate_form3_excel(req, result)
    wb = load_workbook(BytesIO(content))
    assert len(wb.sheetnames) == len(result.листы)
    assert len(result.листы) >= 1


def test_excel_contains_key_fields():
    req = make_form_request()
    result = calculate(req.расчёт)
    content = generate_form3_excel(req, result)
    wb = load_workbook(BytesIO(content))
    ws = wb[wb.sheetnames[0]]
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    text = " ".join(str(v) for v in values)
    assert "Ромашка" in text
    assert "А123ВС777" in text
    assert "Сидоров" in text
    assert "Роснефть" in text  # адрес АЗС попал в маршрут


def test_excel_respects_поля_toggle_off():
    req = make_form_request()
    req.поля.водитель = False
    result = calculate(req.расчёт)
    content = generate_form3_excel(req, result)
    wb = load_workbook(BytesIO(content))
    ws = wb[wb.sheetnames[0]]
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    text = " ".join(str(v) for v in values)
    assert "СНИЛС" not in text


def test_excel_no_lists_still_returns_workbook():
    req = make_form_request(
        расчёт=make_calc_input(водители=[Водитель(фио="А", дни=[])])
    )
    result = calculate(req.расчёт)
    assert result.листы == []
    content = generate_form3_excel(req, result)
    wb = load_workbook(BytesIO(content))
    assert len(wb.sheetnames) == 1  # лист-заглушка "Нет данных"


# ------- PDF -------


def test_pdf_has_one_page_per_list():
    req = make_form_request()
    result = calculate(req.расчёт)
    content = generate_form3_pdf(req, result)
    reader = PdfReader(BytesIO(content))
    assert len(reader.pages) == len(result.листы)


def test_pdf_contains_key_text():
    req = make_form_request()
    result = calculate(req.расчёт)
    content = generate_form3_pdf(req, result)
    reader = PdfReader(BytesIO(content))
    text = "\n".join(p.extract_text() or "" for p in reader.pages)
    assert "ФОРМА" in text.upper() or "ПУТЕВОЙ ЛИСТ" in text.upper()
    assert "А123ВС777" in text


def test_pdf_no_lists_returns_single_placeholder_page():
    req = make_form_request(
        расчёт=make_calc_input(водители=[Водитель(фио="А", дни=[])])
    )
    result = calculate(req.расчёт)
    content = generate_form3_pdf(req, result)
    reader = PdfReader(BytesIO(content))
    assert len(reader.pages) == 1
